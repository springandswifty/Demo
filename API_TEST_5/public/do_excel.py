__author__ = 'zz'
from openpyxl import load_workbook
#处理数据的整体思路：先读初始化数据---处理数据，对相关数据做替换---更新初始化数据
class DoExcel:
    def __init__(self,file_path,sheet_name,mode,case_id_list):
        self.file_path=file_path
        self.sheet_name=sheet_name
        self.mode=mode
        self.case_id_list=case_id_list#传入的是指定测试用例的序号  列表形式的数据传递进来

    def do_excel(self):
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        no_reg_tel=self.get_init_data()#获取初始化手机号？
        mobile=no_reg_tel+1#现有手机号的基础上+1

        test_data=[]#所有行的数据存到这个列表里面
        if self.mode=='1':#如果传进来的mode==1的 就代表要执行所有的用例
            for i in range(2,sheet.max_row+1):# 2 3 4 5
                sub_data={}#每一行的数据单独存在一个字典里面
                sub_data['case_id']=sheet.cell(i,1).value
                sub_data['title']=sheet.cell(i,2).value
                sub_data['method']=sheet.cell(i,3).value
                sub_data['url']=sheet.cell(i,4).value
                sub_data['ExpectedResult']=sheet.cell(i,6).value#添加期望值进来
                if sheet.cell(i,5).value.find('${no_reg_tel}')!=-1:#如果存在这个字符串
                    new_param=sheet.cell(i,5).value.replace('${no_reg_tel}',str(no_reg_tel))
                    sub_data['param']=new_param#为何要保存？
                else:#如果没有${no_reg_tel}  就直接加到字典里面
                    sub_data['param']=sheet.cell(i,5).value
                test_data.append(sub_data)
        else:#如果mode不等于1  那么就取列表里面的值
             for i in eval(self.case_id_list):#加eval str---》list
                sub_data={}#每一行的数据单独存在一个字典里面
                sub_data['case_id']=sheet.cell(i+1,1).value
                sub_data['title']=sheet.cell(i+1,2).value
                sub_data['method']=sheet.cell(i+1,3).value
                sub_data['url']=sheet.cell(i+1,4).value
                sub_data['ExpectedResult']=sheet.cell(i+1,6).value#添加期望值进来
                if sheet.cell(i+1,5).value.find('${no_reg_tel}')!=-1:#如果存在这个字符串
                    new_param=sheet.cell(i+1,5).value.replace('${no_reg_tel}',str(no_reg_tel))
                    sub_data['param']=new_param#为何要保存？
                else:#如果没有${no_reg_tel}  就直接加到字典里面
                    sub_data['param']=sheet.cell(i+1,5).value
                test_data.append(sub_data)
        self.update_init_data(mobile)#更新初始化手机号
        return test_data

    def get_init_data(self):#获取初始化值
        wb=load_workbook(self.file_path)
        sheet=wb['init_data']#可优化
        no_reg_tel=sheet.cell(1,2).value
        return no_reg_tel

    def update_init_data(self,mobile):#更新初始化数据
        wb=load_workbook(self.file_path)
        sheet=wb['init_data']#可优化
        sheet.cell(1,2).value=mobile
        wb.save(self.file_path)

    def write_data(self,r,c,actual_result):#改写函数  指定行列 存储值
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        sheet.cell(r,c).value=actual_result
        wb.save(self.file_path)


if __name__ == '__main__':
    test_data=DoExcel('test_api_1.xlsx','test_data').do_excel()
    #no_reg_tel=DoExcel('test_api_1.xlsx','init_data').get_init_data()
    print(test_data)


#手机号码更新问题的结局
#1:加随机数，存在重复的风险 18688773467 +2  18688773469  +0 18688773469
#2：时间戳 自己去解决
#3：每次注册完了之后 删除这个数据  去数据库里面删除
#4：每次从数据库里面查询最大的手机号  然后在这个基础上+1
#5：自己写一个号码  18688773467 更新Excel里面的号码+1
#6：它必须是一个手机号。